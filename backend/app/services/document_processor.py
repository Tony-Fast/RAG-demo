"""
Document Processing Service
Handles document parsing, text extraction, and chunking
"""

import os
import time
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import hashlib

import aiofiles
from PyPDF2 import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
import charset_normalizer

from app.core.config import settings
from app.models.entities import Document, DocumentChunk, DocumentStatus

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """
    Document processing service for extracting text from various formats
    """
    
    SUPPORTED_FORMATS = {
        '.pdf': 'PDF Document',
        '.docx': 'Word Document',
        '.xlsx': 'Excel Spreadsheet',
        '.xls': 'Excel Spreadsheet',
        '.txt': 'Plain Text',
        '.csv': 'CSV File'
    }
    
    def __init__(self):
        self.chunk_size = settings.CHUNK_SIZE
        self.chunk_overlap = settings.CHUNK_OVERLAP
    
    async def process_document(self, file_path: str, filename: str) -> Tuple[Document, List[DocumentChunk]]:
        """
        Process a document: extract text and create chunks
        
        Args:
            file_path: Path to the uploaded file
            filename: Original filename
            
        Returns:
            Tuple of (Document, List[DocumentChunk])
        """
        start_time = time.time()
        
        # Create document entity
        file_format = Path(filename).suffix.lower()
        file_size = os.path.getsize(file_path)
        
        document = Document.create(
            filename=filename,
            file_path=file_path,
            file_format=file_format,
            file_size=file_size
        )
        
        try:
            # Extract text based on file format
            text = await self._extract_text(file_path, file_format)
            
            # Update document status
            document.status = DocumentStatus.PROCESSING
            document.metadata['text_length'] = len(text)
            document.metadata['extraction_time'] = time.time() - start_time
            
            # Create chunks
            chunks = self._create_chunks(document.id, text)
            document.chunk_count = len(chunks)
            document.status = DocumentStatus.COMPLETED
            
            processing_time = time.time() - start_time
            logger.info(f"Processed document {filename}: {len(text)} chars, {len(chunks)} chunks in {processing_time:.2f}s")
            
            # Create metadata file to store original filename
            metadata_file_path = Path(file_path).with_suffix('.metadata.json')
            metadata = {
                'original_filename': filename,
                'document_id': document.id,
                'processing_time': processing_time,
                'text_length': len(text),
                'chunk_count': len(chunks),
                'created_at': datetime.now().isoformat()
            }
            
            import json
            with open(metadata_file_path, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, ensure_ascii=False, indent=2)
            
            logger.info(f"Created metadata file: {metadata_file_path}")
            
            return document, chunks
            
        except Exception as e:
            document.status = DocumentStatus.FAILED
            document.error_message = str(e)
            logger.error(f"Failed to process document {filename}: {e}")
            raise
    
    async def _extract_text(self, file_path: str, file_format: str) -> str:
        """
        Extract text from document based on format
        
        Args:
            file_path: Path to the file
            file_format: File extension (e.g., .pdf, .docx)
            
        Returns:
            Extracted text content
        """
        text = ""
        
        if file_format == '.pdf':
            text = await self._extract_pdf(file_path)
        elif file_format in ['.docx', '.doc']:
            text = await self._extract_docx(file_path)
        elif file_format in ['.xlsx', '.xls']:
            text = await self._extract_excel(file_path)
        elif file_format == '.csv':
            text = await self._extract_csv(file_path)
        elif file_format == '.txt':
            text = await self._extract_text_file(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_format}")
        
        return text
    
    async def _extract_pdf(self, file_path: str) -> str:
        """
        Extract text from PDF file
        
        Args:
            file_path: Path to PDF file
            
        Returns:
            Extracted text
        """
        try:
            reader = PdfReader(file_path)
            text_parts = []
            
            for page_num, page in enumerate(reader.pages):
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(f"[Page {page_num + 1}]\n{page_text}")
            
            text = "\n\n".join(text_parts)
            
            # Log page count
            logger.info(f"Extracted {len(reader.pages)} pages from PDF")
            
            return text
            
        except Exception as e:
            logger.error(f"PDF extraction failed: {e}")
            raise ValueError(f"Failed to extract text from PDF: {e}")
    
    async def _extract_docx(self, file_path: str) -> str:
        """
        Extract text from Word document
        
        Args:
            file_path: Path to Word file
            
        Returns:
            Extracted text
        """
        try:
            doc = DocxDocument(file_path)
            text_parts = []
            
            current_section = []
            current_heading = ""
            
            for para in doc.paragraphs:
                text = para.text.strip()
                if not text:
                    continue
                
                # Check if paragraph is a heading
                if para.style.name.startswith('Heading'):
                    # Save previous section
                    if current_section:
                        text_parts.append(f"## {current_heading}\n" + "\n".join(current_section))
                    current_heading = text
                    current_section = []
                else:
                    current_section.append(text)
            
            # Save last section
            if current_section:
                text_parts.append(f"## {current_heading}\n" + "\n".join(current_section) if current_heading else "\n".join(current_section))
            
            text = "\n\n".join(text_parts)
            logger.info(f"Extracted {len(text)} characters from DOCX")
            
            return text
            
        except Exception as e:
            logger.error(f"DOCX extraction failed: {e}")
            raise ValueError(f"Failed to extract text from Word document: {e}")
    
    async def _extract_excel(self, file_path: str) -> str:
        """
        Extract text from Excel spreadsheet
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Extracted text
        """
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            text_parts = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"[Sheet: {sheet_name}]")
                
                rows_data = []
                for row in sheet.iter_rows(values_only=True):
                    # Convert row to string representation
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_values):  # Skip empty rows
                        rows_data.append(" | ".join(row_values))
                
                if rows_data:
                    text_parts.append("\n".join(rows_data))
                
                text_parts.append("")  # Empty line between sheets
            
            text = "\n".join(text_parts)
            logger.info(f"Extracted {len(text)} characters from Excel")
            
            return text
            
        except Exception as e:
            logger.error(f"Excel extraction failed: {e}")
            raise ValueError(f"Failed to extract text from Excel file: {e}")
    
    async def _extract_csv(self, file_path: str) -> str:
        """
        Extract text from CSV file
        
        Args:
            file_path: Path to CSV file
            
        Returns:
            Extracted text
        """
        try:
            import csv
            
            text_parts = []
            
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                
                for row_num, row in enumerate(reader):
                    row_values = [cell.strip() for cell in row if cell.strip()]
                    if row_values:
                        if row_num == 0:
                            text_parts.append("Headers: " + " | ".join(row_values))
                        else:
                            text_parts.append(" | ".join(row_values))
            
            text = "\n".join(text_parts)
            logger.info(f"Extracted {len(text)} characters from CSV")
            
            return text
            
        except Exception as e:
            logger.error(f"CSV extraction failed: {e}")
            raise ValueError(f"Failed to extract text from CSV file: {e}")
    
    async def _extract_text_file(self, file_path: str) -> str:
        """
        Extract text from plain text file with encoding detection
        
        Args:
            file_path: Path to text file
            
        Returns:
            Extracted text
        """
        try:
            # Detect encoding
            with open(file_path, 'rb') as f:
                raw_data = f.read()
                result = charset_normalizer.from_bytes(raw_data).best()
                encoding = result.encoding if result else 'utf-8'
            
            # Read with detected encoding
            async with aiofiles.open(file_path, 'r', encoding=encoding) as f:
                text = await f.read()
            
            # Normalize line endings
            text = text.replace('\r\n', '\n').replace('\r', '\n')
            
            logger.info(f"Extracted {len(text)} characters from text file")
            
            return text
            
        except Exception as e:
            logger.error(f"Text file extraction failed: {e}")
            raise ValueError(f"Failed to extract text from file: {e}")
    
    def _create_chunks(self, document_id: str, text: str) -> List[DocumentChunk]:
        """
        Create text chunks from document text
        
        Args:
            document_id: Document ID
            text: Full text content
            
        Returns:
            List of DocumentChunk objects
        """
        chunks = []
        
        if not text or len(text.strip()) == 0:
            return chunks
        
        # Clean and normalize text
        text = self._clean_text(text)
        
        # Split into chunks
        chunk_index = 0
        start = 0
        
        while start < len(text):
            end = start + self.chunk_size
            
            # Try to split at sentence boundary
            if end < len(text):
                # Look for sentence endings near chunk boundary
                sentence_end = self._find_sentence_boundary(text, end)
                if sentence_end > start:
                    end = sentence_end
            
            # Extract chunk
            chunk_text = text[start:end].strip()
            
            if chunk_text:
                # Create metadata
                metadata = {
                    'chunk_char_start': start,
                    'chunk_char_end': end,
                    'chunk_length': len(chunk_text)
                }
                
                chunk = DocumentChunk.create(
                    document_id=document_id,
                    chunk_index=chunk_index,
                    content=chunk_text,
                    metadata=metadata
                )
                chunks.append(chunk)
                chunk_index += 1
            
            # Move start position (with overlap)
            start = end - self.chunk_overlap
            
            # Ensure we make progress
            if start >= len(text):
                break
        
        logger.info(f"Created {len(chunks)} chunks from document")
        
        return chunks
    
    def _clean_text(self, text: str) -> str:
        """
        Clean and normalize text
        
        Args:
            text: Raw text
            
        Returns:
            Cleaned text
        """
        # Remove excessive whitespace
        import re
        
        # Normalize multiple spaces to single space
        text = re.sub(r'[ \t]+', ' ', text)
        
        # Normalize multiple newlines to double newlines
        text = re.sub(r'\n{3,}', '\n\n', text)
        
        # Remove leading/trailing whitespace on each line
        lines = [line.strip() for line in text.split('\n')]
        text = '\n'.join([line for line in lines if line])
        
        return text
    
    def _find_sentence_boundary(self, text: str, position: int) -> int:
        """
        Find a good sentence boundary near the given position
        
        Args:
            text: Full text
            position: Target position
            
        Returns:
            Adjusted position at sentence boundary
        """
        # Look for sentence endings within a window
        window = 200  # Look up to 200 characters ahead
        
        end_pos = min(position + window, len(text))
        
        # Search for sentence endings
        for i in range(position, end_pos):
            if text[i] in '.!?':
                # Check if this looks like a sentence end
                if i + 1 < len(text) and text[i + 1] == ' ':
                    # Skip if followed by lowercase (likely abbreviation)
                    if i + 2 < len(text) and text[i + 2].islower():
                        continue
                    return i + 1
        
        # No good boundary found, return original position
        return position


# Singleton instance
document_processor = DocumentProcessor()
